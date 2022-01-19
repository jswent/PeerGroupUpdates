from openpyxl import load_workbook
from openpyxl.styles import Alignment
import time
import re
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from datetime import date

# -----GLOBAL-----#

dest = "~/Downloads/Peer Group Updates in Q4 2021.xlsx"
dateBound = "August 1, 2021"

# Specify webdriver path and open webpage
driver = webdriver.Chrome(executable_path=r'./chromedriver')
myRow = 3

# Open Excel Spreadsheet and determine which sheet to edit
wb = load_workbook(dest)
sheet = wb.active

def findDateNumber(month):
    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
              'November', 'December']

    for i in range(12):
        if month is months[i]:
            return i + 1

    return None


def checkDate(elementText, dateBoundStr):
    date_split = re.split(' |,', str(elementText))
    bound_split = re.split(' |,', str(dateBoundStr))

    inp_date = date(int(bound_split[2]), findDateNumber(bound_split[1]), int(bound_split[0]))
    pulled_date = date(int(date_split[2]), findDateNumber(date_split[0]), int(date_split[1]))

    if inp_date < pulled_date:
        return True
    else:
        return False


def setWorkbokPath(path):
    dest = path


def setDateBound(dateStr):
    dateBound = dateStr


def runLoop():
    global myRow
    while 0 == 0:

        if len(sheet.cell(row=myRow, column=3).value) > 4 or myRow == 8:
            myRow = myRow + 1

        print(myRow)

        # if myRow == 15 or myRow == 23 or myRow == 17 or myRow == 26 or myRow == 27:
        #     myRow = myRow + 1

        driver.get('https://www.sec.gov/edgar/searchedgar/companysearch.html')

        edgar_search_box = driver.find_element_by_xpath("""//*[@id="company"]""")
        edgar_search_box.send_keys(str(sheet.cell(row=myRow, column=3).value))
        edgar_search_box.send_keys(Keys.ENTER)

        time.sleep(3)

        try:
            select_8k_filings = driver.find_element_by_xpath("""/html/body/main/div[4]/div[2]/div[2]/h5""")
            select_8k_filings.click()
        except:
            myRow = myRow + 1
            continue

        check = False
        index = 1
        events_occurred = ""

        while not check:
            try:
                filing_8k = driver.find_element_by_xpath(
                    """/html/body/main/div[4]/div[2]/div[2]/div/div/ul/li[""" + str(index) + """]/a[1]""")
                if checkDate(filing_8k.text):
                    events_occurred_local = driver.find_element_by_xpath(
                        """/html/body/main/div[4]/div[2]/div[2]/div/div/ul/li[""" + str(index) + """]/small""")
                    events_occurred = events_occurred + events_occurred_local.text + "\n"
                else:
                    check = True
                index = index + 1
            except:
                check = True

        sheet.cell(row=myRow, column=5).value = events_occurred
        sheet.cell(row=myRow, column=5).alignment = Alignment(wrapText=True)

        most_recent_filing = driver.find_element_by_xpath(
            """/html/body/main/div[4]/div[2]/div[2]/div/div/ul/li[1]/a[1]""")
        filing_date = most_recent_filing.text
        # print(filing_date)
        date_split = re.split(' |,', str(filing_date))

        if (date_split[0] == 'May' and int(date_split[1]) >= 15) or date_split[0] == 'June' or date_split[0] == 'July':
            sheet.cell(row=myRow, column=4).value = "Yes"
        else:
            sheet.cell(row=myRow, column=4).value = "No"

        # events_occurred = driver.find_element_by_xpath("""/html/body/main/div[4]/div[2]/div[2]/div/div/ul/li[1]/small/ul""")
        # sheet.cell(row=myRow,column=5).value = str(events_occurred.text)
        # sheet.cell(row=myRow,column=5).alignment = Alignment(wrapText=True)

        # ---Check 10Q---#
        select_10kq_filings = driver.find_element_by_xpath("""/html/body/main/div[4]/div[2]/div[3]/h5""")
        select_10kq_filings.click()

        most_recent_filing = driver.find_element_by_xpath(
            """/html/body/main/div[4]/div[2]/div[2]/div/div/ul/li[1]/a[1]""")
        filing_date = most_recent_filing.text
        # print(sheet.cell(row=myRow,column=2).value + " Filing Date: " +  filing_date)
        date_split = re.split(' |,', str(filing_date))

        if (date_split[0] == 'May' and int(date_split[1]) >= 15) or date_split[0] == 'June' or date_split[0] == 'July':
            sheet.cell(row=myRow, column=6).value = "Yes"
        else:
            # print(date_split[0] + " | " + date_split[1])
            sheet.cell(row=myRow, column=6).value = "No"
        # ---END---#

        if "5.07" in sheet.cell(row=myRow, column=5).value:
            check = False
            index = 1

            while not check:
                filing_date_scan = driver.find_element_by_xpath(
                    """/html/body/main/div[4]/div[2]/div[2]/div/div/ul/li[""" + str(index) + """]/a[1]""")
                scan_events_occurred = driver.find_element_by_xpath(
                    """/html/body/main/div[4]/div[2]/div[2]/div/div/ul/li[""" + str(index) + """]/small""")
                if "5.07" in scan_events_occurred.text:
                    filing_date_scan.click()

        wb.save(dest)
        myRow = myRow + 1

# --------------------------#
